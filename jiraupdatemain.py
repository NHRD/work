import pandas as pd
import requests
import sys
import logging
import os
import openpyxl
from jira_auth import get_auth_info
from openpyxl import load_workbook, styles
from shutil import copy

# ロギング設定
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler('jira_sync.log', encoding='utf-8')
    ]
)
logger = logging.getLogger(__name__)

# URLからチケットキーを抽出
# 例: https://jira.example.com/browse/ABC-123 → ABC-123
def extract_key(url):
    return url.split('/')[-1] if pd.notna(url) and url else None

# Summaryの先頭にプロジェクト名のヘッダーを追加（重複防止）
def ensure_summary_header(summary, project_label):
    header = f"[{project_label}]"
    if summary.startswith(header):
        return summary
    return f"{header} {summary}"

# JIRAのコメントAPIからチケットの最後のコメントを取得
# Excelへの書き戻し用に使用
def get_last_comment(jira_url, key, headers):
    try:
        comments_url = f"{jira_url}/rest/api/2/issue/{key}/comment"
        resp = requests.get(comments_url, headers=headers)
        resp.raise_for_status()
        comments = resp.json().get("comments", [])
        return comments[-1]["body"] if comments else ""
    except Exception as e:
        logger.error(f"コメント取得失敗 ({key}): {str(e)}")
        return ""

# JIRAのチケットからAssigneeの表示名を取得してExcelへ転記する
# "山田 太郎 (uid12345)"の形式で返す
def get_jira_assignee_name(fields):
    assignee_info = fields.get("assignee")
    if assignee_info and assignee_info.get("displayName"):
        return assignee_info["displayName"] + f" ({assignee_info.get('name', '')})"
    return ""

# JIRAに新規チケットを作成する
def create_jira_ticket(jira_url, headers, project_key, row, idx, df):
    summary_value = str(row.get("Summary", "")).strip()
    summary = ensure_summary_header(summary_value, project_key)
    description_raw = row.get("Description", "")
    description = str(description_raw) if pd.notna(description_raw) else ""
    due_date_raw = row.get("Due Date", "")
    due_date_str = str(due_date_raw) if pd.notna(due_date_raw) else None
    comment = str(row.get("Comment", "")) if pd.notna(row.get("Comment")) else None

    logger.info("+ 新規作成 (Subaru)")
    
    try:
        # JIRAの課題作成APIへ送信
        payload = {
            "fields": {
                "project": {"key": project_key},
                "summary": summary,
                "description": description,
                "duedate": due_date_str,
                "issuetype": {"name": "Task"},
                "labels": ["Customer_QA"],
                "assignee": {"name": "uig17323"}  # APIユーザーに自動アサイン
            }
        }
        
        res = requests.post(f"{jira_url}/rest/api/2/issue", headers=headers, json=payload)
        res.raise_for_status()
        
        new_key = res.json()["key"]
        new_url = f"{jira_url}/browse/{new_key}"
        df.at[idx, "Ticket URL"] = new_url
        df.at[idx, "Ticket Key"] = new_key
        logger.info(f"✓ 作成: {new_key}")
        
        # コメントがあれば追加
        if comment:
            comment_url = f"{jira_url}/rest/api/2/issue/{new_key}/comment"
            comment_res = requests.post(comment_url, headers=headers, json={"body": comment})
            comment_res.raise_for_status()
            
        df.at[idx, "Sync"] = ""
        df.at[idx, "Assignee"] = "Harada, Naohisa (uig17323)"
        return True, new_key
        
    except Exception as e:
        logger.error(f"x 作成失敗: {str(e)}")
        return False, None

# 既存のJIRAチケットを更新する
def update_jira_ticket(jira_url, headers, project_key, row, idx, df, key):
    summary_value = str(row.get("Summary", "")).strip()
    summary = ensure_summary_header(summary_value, project_key)
    description_raw = row.get("Description", "")
    description = str(description_raw) if pd.notna(description_raw) else ""
    due_date_raw = row.get("Due Date", "")
    due_date_str = str(due_date_raw) if pd.notna(due_date_raw) else None
    comment = str(row.get("Comment", "")) if pd.notna(row.get("Comment")) else None

    logger.info(f".. 更新: {key} (Subaru)")
    
    try:
        # JIRA課題更新APIでサマリ・説明・期限を更新
        update_url = f"{jira_url}/rest/api/2/issue/{key}"
        payload = {
            "fields": {
                "summary": summary,
                "description": description,
                "duedate": due_date_str
            }
        }
        
        res = requests.put(update_url, headers=headers, json=payload)
        res.raise_for_status()
        logger.info(f"✓ 更新成功: {key}")
        
        # コメントがあれば追加
        if comment:
            # 現在のコメントを取得して比較
            current_comment = get_last_comment(jira_url, key, headers)
            if current_comment != comment:
                comment_url = f"{jira_url}/rest/api/2/issue/{key}/comment"
                comment_res = requests.post(comment_url, headers=headers, json={"body": comment})
                comment_res.raise_for_status()
                logger.info(f"✓ コメント追加: {key}")
            
        df.at[idx, "Sync"] = ""

        # Assignee を JIRA から取得し Excel に反映（確実な一致を保証）
        issue_res = requests.get(f"{jira_url}/rest/api/2/issue/{key}?fields=assignee", headers=headers)
        issue_res.raise_for_status()
        assignee_name = get_jira_assignee_name(issue_res.json().get("fields", {}))
        if assignee_name:
            df.at[idx, "Assignee"] = assignee_name
            
        return True
        
    except Exception as e:
        logger.error(f"! 更新失敗: {str(e)}")
        return False

# JIRAからチケットを検索し、Excelに存在しないものを追加
def import_jira_tickets(jira_url, headers, project_key, df):
    logger.info("JIRAからチケットをインポート開始")
    added_count = 0
    updated_count = 0
    
    try:
        # JQLクエリでCustomer_QAラベルかつDone/CANCELED以外のチケットを検索
        jql = f'project = {project_key} AND labels = Customer_QA AND status NOT IN (Done, CANCELED)'
        res = requests.get(f"{jira_url}/rest/api/2/search?jql={jql}&maxResults=1000", headers=headers)
        res.raise_for_status()
        
        issues = res.json().get("issues", [])
        logger.info(f"JIRAから取得したチケット数: {len(issues)}")
        
        # 現在のExcelのURLからキーのリストを作成
        existing_keys = {}  # キーとExcelの行インデックスのマッピング
        for idx, row in df.iterrows():
            url = row.get("Ticket URL", "")
            if pd.notna(url) and url:
                key = url.split('/')[-1]
                existing_keys[key] = idx
        
        for issue in issues:
            key = issue["key"]
            url = f"{jira_url}/browse/{key}"
            
            # Excel内に存在するかチェック
            if key not in existing_keys:
                # 新規追加
                fields = issue["fields"]
                summary = fields.get("summary", "")
                description = fields.get("description", "") or ""
                due_date = fields.get("duedate", "")
                comment = get_last_comment(jira_url, key, headers)
                assignee_name = get_jira_assignee_name(fields)
                
                # 次のNo.値を取得
                next_no = df["No."].dropna().max() + 1 if not df["No."].dropna().empty else 1
                
                # 新規行をDataFrameに追加
                new_row = pd.DataFrame.from_records([{
                    "No.": next_no,
                    "Ticket URL": url,
                    "Summary": summary,
                    "Assignee": assignee_name or "Subaru",  # アサイニーがない場合はSubaruをデフォルト値に
                    "Description": description,
                    "Due Date": due_date,
                    "Comment": comment,
                    "Sync": "",
                    "Status": ""
                }])
                
                df = pd.concat([df, new_row], ignore_index=True)
                added_count += 1
                logger.info(f"Excelに追加: {key} - {summary}")
            else:
                # 既存のチケットの場合、コメントとアサイニーを更新
                idx = existing_keys[key]
                
                # アサイニーの確認と更新
                fields = issue["fields"]
                assignee_name = get_jira_assignee_name(fields)
                current_assignee = df.at[idx, "Assignee"] if pd.notna(df.at[idx, "Assignee"]) else ""
                is_subaru = current_assignee.strip().lower() == "subaru"
                
                # ステータスがDoneでない場合のみ処理
                status = str(df.at[idx, "Status"]).strip().lower()
                if status != "done":
                    # コメントの更新
                    jira_comment = get_last_comment(jira_url, key, headers)
                    excel_comment = df.at[idx, "Comment"] if pd.notna(df.at[idx, "Comment"]) else ""
                    
                    # アサイニーがSubaruの場合
                    if is_subaru:
                        # Syncがあるかどうかをチェック
                        sync_value = str(df.at[idx, "Sync"]).strip()
                        
                        if sync_value == "〇":
                            # Syncがある場合のみJIRAの値で更新
                            if assignee_name:
                                df.at[idx, "Assignee"] = assignee_name
                                logger.info(f"アサイニー更新: {key} - {assignee_name}")
                                updated_count += 1
                            
                            # コメントも更新
                            if jira_comment and jira_comment != excel_comment:
                                df.at[idx, "Comment"] = jira_comment
                                logger.info(f"コメント更新 (Subaru担当): {key}")
                                updated_count += 1
                        else:
                            logger.info(f"スキップ (Syncなし、Subaru担当): {key}")
                            # Syncがない場合は更新しない
                    
                    # アサイニーがSubaruでない場合、コメントの確認・更新とアサイニーをSubaruに変更
                    elif not is_subaru:
                        # JIRAの最新コメントとExcelのコメントが異なる場合更新
                        if jira_comment and jira_comment != excel_comment:
                            df.at[idx, "Comment"] = jira_comment
                            df.at[idx, "Assignee"] = "Subaru"
                            logger.info(f"コメント更新とSubaruへ担当変更: {key}")
                            updated_count += 1
                        
        logger.info(f"JIRAからの更新完了。追加: {added_count}件、更新: {updated_count}件")
        return df, added_count, updated_count
        
    except Exception as e:
        logger.error(f"JIRAからのインポート・更新失敗: {str(e)}")
        return df, 0, 0

# 定数定義
MAIN_SHEET_NAME = "main"  # メインシートの名前
IMAGES_SHEET_NAME = "Imanges"  # 画像シートの名前（表記ゆれに対応）

# Excelの書式を設定する（メインシートのみ）
def format_excel_file(excel_path):
    try:
        wb = load_workbook(excel_path)
        
        # メインシートを対象とする
        if MAIN_SHEET_NAME in wb.sheetnames:
            ws = wb[MAIN_SHEET_NAME]
        else:
            # メインシート名が存在しない場合は最初のシートを使用
            ws = wb.active
            logger.warning(f"シート '{MAIN_SHEET_NAME}' が見つからないため、最初のシート '{ws.title}' に書式を適用します")
        
        # 薄緑の背景色設定
        fill = styles.PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
        
        # 罫線設定
        border = styles.Border(
            left=styles.Side(style='thin'),
            right=styles.Side(style='thin'),
            top=styles.Side(style='thin'),
            bottom=styles.Side(style='thin')
        )
        
        # 全セルに罫線を設定し、偶数行に背景色を設定（メインシートのみ）
        for row in range(1, ws.max_row + 1):
            for col in range(1, 10):  # A-I列
                cell = ws.cell(row=row, column=col)
                cell.border = border
                if row % 2 == 0:  # 偶数行
                    cell.fill = fill
        
        # 他のシートには書式設定を適用しない
        logger.info(f"シート '{ws.title}' のみに書式設定を適用しました")
                    
        wb.save(excel_path)
        logger.info("✓ Excel書式設定完了")
        return True
        
    except Exception as e:
        logger.error(f"Excel書式設定失敗: {str(e)}")
        return False

# ExcelとJIRAを同期する主関数
def sync_excel_and_jira(excel_path, project_key):
    logger.info("ExcelとJIRAの同期処理を開始します")
    
    try:
        # JIRA認証情報取得
        token, jira_url, headers = get_auth_info()
        
        # Excelファイルが存在するか確認
        file_exists = os.path.isfile(excel_path)
        
        # 既存のファイルからImagesシートなどを保存（ファイルがある場合）
        other_sheets = {}
        if file_exists:
            try:
                existing_wb = load_workbook(excel_path)
                # 最初のシート（メインシート）以外のシートを保存
                main_sheet_name = existing_wb.sheetnames[0]
                for sheet_name in existing_wb.sheetnames:
                    if sheet_name != main_sheet_name:
                        other_sheets[sheet_name] = existing_wb[sheet_name]
                logger.info(f"保存する追加シート: {', '.join(other_sheets.keys()) if other_sheets else 'なし'}")
            except Exception as e:
                logger.error(f"既存シートの読み込みエラー: {str(e)}")
        
        # Excelファイルを読み込み。ファイルがなければ空のテンプレートを作成
        try:
            df = pd.read_excel(excel_path)
        except FileNotFoundError:
            df = pd.DataFrame(columns=["No.", "Ticket URL", "Summary", "Assignee", "Description", "Due Date", "Comment"])
        
        # 欠けている列の初期化
        if "Status" not in df.columns:
            df["Status"] = ""
        if "Sync" not in df.columns:
            df.insert(df.columns.get_loc("Comment") + 1, "Sync", "〇")
        if "Ticket URL" not in df.columns:
            df["Ticket URL"] = ""
        
        # URL列からキー列を追加（処理用）
        df["Ticket Key"] = df["Ticket URL"].apply(extract_key)
        
        # 処理統計用変数
        created_keys = set()
        updated_keys = set()
        
        # Excel行ごとの処理
        if "No." in df.columns:
            if df["No."].dropna().empty:
                max_no = 0
            else:
                max_no = int(df["No."].dropna().max())
                
            for no in range(1, max_no + 1):
                matched = df[df["No."] == no]
                if matched.empty:
                    continue
                    
                idx = matched.index[0]
                row = df.loc[idx]
                
                # Doneチケットはスキップ
                if str(row.get("Status", "")).strip().lower() == "done":
                    continue
                
                # Syncが〇でないものは対象外
                if str(row.get("Sync", "")).strip() != "〇":
                    continue
                
                summary_value = str(row.get("Summary", "")).strip()
                key = row.get("Ticket Key")
                
                # チケットURLがない → 新規作成処理
                if summary_value and (not key or key.strip() == ""):
                    success, new_key = create_jira_ticket(jira_url, headers, project_key, row, idx, df)
                    if success and new_key:
                        created_keys.add(new_key)
                
                # 既存のチケットがある → 更新処理
                elif key and key.strip():
                    assignee = row["Assignee"] if pd.notna(row["Assignee"]) else ""
                    is_subaru = assignee.strip().lower() == "subaru"
                    sync_value = str(row.get("Sync", "")).strip()
                    
                    # 担当者がSubaruで、Syncが〇のもののみ更新対象
                    if not is_subaru:
                        logger.info(f"スキップ (Subaru以外): {key}")
                        continue
                    
                    # Syncが〇の場合のみ更新
                    if sync_value == "〇":
                        success = update_jira_ticket(jira_url, headers, project_key, row, idx, df, key)
                        if success:
                            updated_keys.add(key)
                    else:
                        logger.info(f"スキップ (Syncなし): {key}")
                        continue
        
        # JIRAからチケットをインポートして既存チケットも更新
        df, added_count, updated_count = import_jira_tickets(jira_url, headers, project_key, df)
        
        # 統計情報をログ出力
        logger.info("=" * 30)
        logger.info(f"同期処理結果サマリー:")
        logger.info(f"  作成したチケット数: {len(created_keys)}")
        logger.info(f"  更新したチケット数: {len(updated_keys)}")
        logger.info(f"  インポートしたチケット数: {added_count}")
        logger.info(f"  JIRAからの更新数: {updated_count}")
        
        # 一時列削除
        df.drop(columns=["Ticket Key"], inplace=True)

        # 既存のExcelファイルが存在する場合、ハイパーリンクとシートを保持した更新を行う
        if file_exists:
            # 一時的な新しいファイル名（拡張子をexcelと同じにする）
            file_path_without_ext = os.path.splitext(excel_path)[0]
            temp_file = file_path_without_ext + "_tmp.xlsx"
            backup_file = file_path_without_ext + "_backup.xlsx"
            
            # バックアップとして元のファイルをコピー（画像保持のため）
            copy(excel_path, backup_file)
            
            # まず新しいDataFrameを一時ファイルに保存
            df.to_excel(temp_file, index=False)
            
            # 既存のワークブックを読み込む
            existing_wb = load_workbook(excel_path)
            existing_sheet = existing_wb.active
            
            # 新しいワークブックを読み込む
            new_wb = load_workbook(temp_file)
            new_sheet = new_wb.active
            
            # J列のセルをコピー（ハイパーリンクを保持）
            for row in range(1, existing_sheet.max_row + 1):
                # J列にデータがある場合
                if row <= existing_sheet.max_row and existing_sheet.cell(row=row, column=10).value is not None:
                    # 対応する行が新しいシートに存在するかチェック
                    if row <= new_sheet.max_row:
                        # J列のセルをコピー（ハイパーリンク情報を含む）
                        new_sheet.cell(row=row, column=10).value = existing_sheet.cell(row=row, column=10).value
                        # ハイパーリンクがある場合それも保持
                        if existing_sheet.cell(row=row, column=10).hyperlink:
                            new_sheet.cell(row=row, column=10).hyperlink = existing_sheet.cell(row=row, column=10).hyperlink
            
            # 更新された内容を元のファイル名で保存
            new_wb.save(excel_path)
            
            # 画像を含むシートがある場合は特別な処理
            if any("Imanges" in sheet_name for sheet_name in other_sheets.keys()):
                logger.info("画像を含むシートを検出しました。元のファイルを保持します。")
                try:
                    # バックアップから元のワークブックを読み込む
                    original_wb = load_workbook(backup_file)
                    # 新しく作成したワークブックを読み込む
                    updated_wb = load_workbook(excel_path)
                    
                    # 最初のシート（メインシート）のデータを更新
                    new_main_sheet = updated_wb.active
                    original_main_sheet = original_wb.active
                    
                    # メインシートを削除して新しいものに置き換え
                    main_sheet_name = original_wb.sheetnames[0]
                    del original_wb[main_sheet_name]
                    
                    # シートをコピー
                    original_wb.create_sheet(main_sheet_name, 0)
                    replaced_sheet = original_wb[main_sheet_name]
                    
                    # セルの内容をコピー
                    for row in range(1, new_main_sheet.max_row + 1):
                        for col in range(1, new_main_sheet.max_column + 1):
                            replaced_sheet.cell(row=row, column=col).value = new_main_sheet.cell(row=row, column=col).value
                            
                            # ハイパーリンクもコピー
                            if new_main_sheet.cell(row=row, column=col).hyperlink:
                                replaced_sheet.cell(row=row, column=col).hyperlink = new_main_sheet.cell(row=row, column=col).hyperlink
                                
                            # スタイルもコピー
                            try:
                                if new_main_sheet.cell(row=row, column=col).has_style:
                                    replaced_sheet.cell(row=row, column=col).font = new_main_sheet.cell(row=row, column=col).font
                                    replaced_sheet.cell(row=row, column=col).border = new_main_sheet.cell(row=row, column=col).border
                                    replaced_sheet.cell(row=row, column=col).fill = new_main_sheet.cell(row=row, column=col).fill
                                    replaced_sheet.cell(row=row, column=col).number_format = new_main_sheet.cell(row=row, column=col).number_format
                                    replaced_sheet.cell(row=row, column=col).alignment = new_main_sheet.cell(row=row, column=col).alignment
                            except Exception as style_error:
                                logger.warning(f"メインシートスタイルのコピー中にエラー: {style_error}")
                    
                    # 元のファイルに保存
                    original_wb.save(excel_path)
                    logger.info(f"画像を保持したまま、メインシートを更新しました")
                    
                except Exception as e:
                    logger.error(f"画像保持処理エラー: {str(e)}")
                    logger.info("通常の保存処理にフォールバックします")
                    
                    # 通常の保存処理（画像は失われるがデータは保持される）
                    df.to_excel(excel_path, index=False)
                    format_excel_file(excel_path)
            else:
                # 保存していた他のシートを新しいワークブックに追加
                for sheet_name, sheet in other_sheets.items():
                    # すでに同名のシートがある場合は削除
                    if sheet_name in new_wb.sheetnames:
                        std = new_wb[sheet_name]
                        new_wb.remove(std)
                    
                    # シートを新しいワークブックにコピー
                    new_sheet = new_wb.create_sheet(title=sheet_name)
                    
                    # シートの内容をコピー
                    for row_idx, row in enumerate(sheet.rows, 1):
                        for col_idx, cell in enumerate(row, 1):
                            new_cell = new_sheet.cell(row=row_idx, column=col_idx)
                            # セルの値をコピー
                            new_cell.value = cell.value
                            # ハイパーリンクをコピー
                            if cell.hyperlink:
                                new_cell.hyperlink = cell.hyperlink
                            
                            # スタイルのコピーは個別に行う（styleproxyはハッシュ化できないため）
                            try:
                                # フォント
                                if cell.font:
                                    new_cell.font = styles.Font(
                                        name=cell.font.name,
                                        size=cell.font.size,
                                        bold=cell.font.bold,
                                        italic=cell.font.italic,
                                        vertAlign=cell.font.vertAlign,
                                        underline=cell.font.underline,
                                        strike=cell.font.strike,
                                        color=cell.font.color
                                    )
                                # 罫線
                                if cell.border:
                                    new_cell.border = styles.Border(
                                        left=styles.Side(style=cell.border.left.style, color=cell.border.left.color) if cell.border.left else None,
                                        right=styles.Side(style=cell.border.right.style, color=cell.border.right.color) if cell.border.right else None,
                                        top=styles.Side(style=cell.border.top.style, color=cell.border.top.color) if cell.border.top else None,
                                        bottom=styles.Side(style=cell.border.bottom.style, color=cell.border.bottom.color) if cell.border.bottom else None
                                    )
                                # 塗りつぶし
                                if cell.fill:
                                    new_cell.fill = styles.PatternFill(
                                        fill_type=cell.fill.fill_type,
                                        start_color=cell.fill.start_color,
                                        end_color=cell.fill.end_color
                                    )
                                # 数値書式
                                new_cell.number_format = cell.number_format
                                # 配置
                                if cell.alignment:
                                    new_cell.alignment = cell.alignment
                            except Exception as style_error:
                                logger.warning(f"セルスタイルのコピー中にエラー: {style_error}")
                                # スタイルコピーのエラーは無視して続行
                    
                    # 列の幅をコピー
                    for col_idx, col in enumerate(sheet.columns, 1):
                        letter = openpyxl.utils.get_column_letter(col_idx)
                        if sheet.column_dimensions[letter].width is not None:
                            new_sheet.column_dimensions[letter].width = sheet.column_dimensions[letter].width
                    
                    # 行の高さをコピー
                    for row_idx in range(1, sheet.max_row + 1):
                        if sheet.row_dimensions[row_idx].height is not None:
                            new_sheet.row_dimensions[row_idx].height = sheet.row_dimensions[row_idx].height
                
                # 更新された内容を元のファイル名で保存
                new_wb.save(excel_path)
            
            # 一時ファイルとバックアップを削除
            try:
                os.remove(temp_file)
                os.remove(backup_file)
            except Exception as e:
                logger.warning(f"一時ファイル削除エラー: {str(e)}")
            
            # 書式の調整（メインシートのみ）
            format_excel_file(excel_path)
        else:
            # 新規ファイルの場合は通常の保存
            df.to_excel(excel_path, index=False)
            format_excel_file(excel_path)
        
        logger.info("✓ Excel保存完了（同期処理）")
        return True
        
    except Exception as e:
        logger.error(f"同期処理エラー: {str(e)}")
        return False

# メイン関数
def main():
    if len(sys.argv) < 3:
        print("使い方: python main.py [Excelファイルパス] [JIRAプロジェクトキー]")
        sys.exit(1)

    _, excel_path, project_key = sys.argv
    success = sync_excel_and_jira(excel_path, project_key)
    
    if not success:
        logger.error("同期処理が失敗しました")
        sys.exit(1)
    else:
        logger.info("同期処理が正常に完了しました")

if __name__ == "__main__":
    main()
